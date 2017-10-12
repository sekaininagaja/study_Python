#! python34
# updateProduce.py - 生産販売スプレッドシートのコストを修正します。

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# 生産タイプとその更新価格
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# 行をループして価格を更新します。
for row_num in range(2, sheet.max_row):  # skip the first row
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

wb.save('updateProduceSales.xlsx')
